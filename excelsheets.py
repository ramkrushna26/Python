
import openpyxl, os

workbook = openpyxl.load_workbook('<your excel>')
print(type(workbook))

#get sheet by name
sheet = workbook.get_sheet_by_name('<sheet name>')
print(type(sheet))

#get sheet names
print(workbook.get_sheet_names())

#get the value from sheet
cell = sheet['A1']
#or
cell = sheet.cell(row=1, column=1)
print(cell.value)
print(type(cell))

#creating new workbook
wb = openpyxl.Workbook()
print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('<sheet name>')

#assigning value to cell
sheet['A1'] = 32

#to create new sheet
nsheet = wb.create_sheet("<sheet name>")
#to rename sheet
nsheet.title = "sheetname"
print(wb.get_sheet_names())

#create sheet at specific index
msheet = wb.create_Sheet(index = 0, "sheet name")

#to save to excel
wb.save('<file>.xlsx')
